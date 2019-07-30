import openpyxl
import time
import warnings
warnings.filterwarnings("ignore")

class ExcelFile:
    def __init__(self,excelFilePath):
        self.excelFilePath = excelFilePath
        self.workBook = self.load(excelFilePath)

    def load(self,excelFilePath):
        try:
            workBook = openpyxl.load_workbook(excelFilePath)
        except:
            workBook = openpyxl.Workbook()
        return workBook

    def insert(self,dataList,sheetName):
        if self.workBook is None:
            self.workBook = self.load(self.excelFilePath)
        try:
            book = self.workBook.get_sheet_by_name(sheetName)
        except:
            book = self.workBook.create_sheet(sheetName)
        book.append(dataList)

    def save(self):
        if self.workBook is not None:
            self.workBook.save(self.excelFilePath)

    def remove(self,sheetName):
        if self.workBook is not None:
            book = self.workBook.get_sheet_by_name(sheetName)
            self.workBook.remove_sheet(book)

if __name__ == "__main__":
    excel = ExcelFile("D:/yarnmonitor/log/application_1557411283892_0026.xlsx")
    index = 0
    excel.insert([1, 2, 3, index], "completed_job")
    excel.insert([1, 2, 3, index], "completed_stage")
    excel.remove("Sheet")
    while True:
        excel.insert([1,2,3,index],"completed_job")
        excel.insert([2,3,4,index],"completed_job")
        excel.insert([2, 3, 4, index], "completed_stage")
        excel.save()
        index += 1
        time.sleep(1)
