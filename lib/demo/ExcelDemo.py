# encoding:utf-8
import sys
import os
import openpyxl
sys.path.append(os.path.abspath("../.."))
import lib.util.YmlUtil as ymlUtil
import warnings
warnings.filterwarnings("ignore")
yml = ymlUtil.loadYaml(os.path.abspath("../..") + "/etc/yarnmonitor.yml")
rootPath = yml.get("log_root_path")
workbook = openpyxl.load_workbook(rootPath + '/new_excel.xlsx')
sheetnames = workbook.get_sheet_names()
sheet = workbook.get_sheet_by_name(sheetnames[0])
nrows = sheet.max_row # 获得行数
ncolumns = sheet.max_column # 获得行数
values = ['E','X','C','E','L']
ws = workbook.active
ws.append([1,2,3])
workbook.save(rootPath + '/new_excel.xlsx')